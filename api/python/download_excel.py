
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
import requests
import urllib3
import time
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from fastapi.middleware.cors import CORSMiddleware

# Suppress InsecureRequestWarning
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = FastAPI()

# CORS 미들웨어 추가
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

proxies = {
    'http': None,
    'https': None,
}

TRADE_TYPE_MAPPING = {
    "전체": "",
    "매매": "A1",
    "전세": "B1",
    "월세": "B2"
}

# --- 데이터 수집 함수들 (fetch_listings.py와 중복되지만, 서버리스 함수는 독립적으로 실행되므로 포함) ---

def get_real_estate_data(complex_no: str, trade_type: str, page: int = 1):
    try:
        cookies = {
            'NNB': 'FGYNFS4Y6M6WO',
            'REALESTATE': 'Tue Jan 28 2025 16:23:02 GMT+0900 (Korean Standard Time)',
            'NFS': '2',
            'ASID': 'afd10077000001934e8033f50000004e',
            'ba.uuid': 'a5e52e8f-1775-4eea-9b42-30223205f9df',
        }
        headers = {
            'accept': '*/*',
            'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZCI6IlJFQUxFU1RBVEUiLCJpYXQiOjE3MzgwNDcxNjMsImV4cCI6MTczODA1Nzk2M30.Heq-J33LY9pJDnYOqmRhTTrSPqCpChtWxka_XUphnd4',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
            'referer': f'https://new.land.naver.com/complexes/{complex_no}',
        }
        params = {
            'realEstateType': 'APT:PRE:ABYG:JGC',
            'tradeType': trade_type,
            'tag': '::::::::',
            'rentPriceMin': '0', 'rentPriceMax': '900000000',
            'priceMin': '0', 'priceMax': '900000000',
            'areaMin': '0', 'areaMax': '900000000',
            'showArticle': 'false', 'sameAddressGroup': 'false',
            'priceType': 'RETAIL', 'page': str(page),
            'complexNo': str(complex_no), 'type': 'list', 'order': 'rank'
        }
        url = f'https://new.land.naver.com/api/articles/complex/{complex_no}'
        
        session = requests.Session()
        session.headers.update(headers)
        session.cookies.update(cookies)
        session.proxies.update(proxies)

        # 세션 초기화
        init_url = f'https://new.land.naver.com/complexes/{complex_no}'
        init_response = session.get(init_url, verify=False, timeout=10)
        init_response.raise_for_status()

        response = session.get(url, params=params, timeout=30, verify=False)
        response.raise_for_status()
        return response.json()
    except Exception:
        return None

def process_data(data):
    if not data or 'articleList' not in data:
        return []
    articles = data.get('articleList', [])
    processed_data = []
    for article in articles:
        tags = ', '.join(article.get('tagList', [])) if article.get('tagList') else ''
        processed_article = {
            '순번': 'N/A',
            '아파트명': article.get('articleName'),
            '거래유형': article.get('tradeTypeName'),
            '층수': article.get('floorInfo'),
            '월세': article.get('rentPrc'),
            '거래가격': article.get('dealOrWarrantPrc'),
            '면적(m²)': article.get('area2'),
            '방향': article.get('direction'),
            '등록일': article.get('articleConfirmYmd'),
            '동': article.get('buildingName'),
            '중개사무소': article.get('realtorName'),
            '특징': tags
        }
        processed_data.append(processed_article)
    return processed_data

def fetch_all_pages(complex_no: str, trade_type_key: str):
    all_data = []
    page = 1
    trade_type = TRADE_TYPE_MAPPING.get(trade_type_key, "")
    while True:
        response_data = get_real_estate_data(complex_no, trade_type, page)
        if not response_data or not response_data.get('articleList'): break
        all_data.extend(process_data(response_data))
        if not response_data.get('isMoreData', False): break
        page += 1
        time.sleep(0.5)
    all_data.sort(key=lambda x: x['등록일'] if x['등록일'] else '99999999')
    return all_data

# --- 엑셀 생성 함수 ---
def create_excel_in_memory(data):
    if not data:
        return None
    
    title = data[0].get('아파트명', '부동산_데이터')
    # 엑셀 데이터에는 아파트명 열이 필요 없으므로 제거
    for d in data:
        if '아파트명' in d:
            del d['아파트명']
            
    headers = list(data[0].keys())
    num_columns = len(headers)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title if len(title) <= 31 else title[:31]
    ws.sheet_view.showGridLines = False

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=18, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = light_green_fill

    for row_idx, row_data in enumerate(data, start=4):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if header == '순번':
                cell.value = row_idx - 3
            else:
                cell.value = row_data.get(header)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook, f"{title}.xlsx"

@app.get("/")
async def download_excel_api(complex_no: str, trade_type: str):
    if not complex_no or not trade_type:
        raise HTTPException(status_code=400, detail="complex_no와 trade_type을 모두 입력해주세요.")
    
    try:
        all_data = fetch_all_pages(complex_no, trade_type)
        if not all_data:
            raise HTTPException(status_code=404, detail="다운로드할 데이터가 없습니다.")
        
        virtual_workbook, filename = create_excel_in_memory(all_data)
        
        return StreamingResponse(virtual_workbook, 
                                 media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                 headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"엑셀 생성 중 오류 발생: {e}")
